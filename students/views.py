print('students views.py loaded')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML
import tempfile
import openpyxl
from django.db.models import Q
from django.contrib import messages
from io import BytesIO
from datetime import datetime

from .models import Student, ClassSection, AcademicYear
from .forms import StudentForm

# Add Student
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML
import tempfile
import openpyxl
from django.db.models import Q
from django.contrib import messages
from io import BytesIO
from datetime import datetime

from .models import Student, ClassSection, AcademicYear
from .forms import StudentForm

# Add Student
@login_required
def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                student = form.save(commit=False)
                # Ensure required fields are set
                if not student.date_joined:
                    student.date_joined = datetime.now().date()
                student.save()
                messages.success(request, 'Student added successfully!')
                return redirect('student_list')
            except Exception as e:
                messages.error(request, f'Error saving student: {str(e)}')
        else:
            # Print form errors to console for debugging
            print("Form errors:", form.errors)
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm()
    
    # Get all class sections and academic years for the template
    class_sections = ClassSection.objects.all()
    academic_years = AcademicYear.objects.all()
    
    return render(request, 'student_add.html', {
        'form': form,
        'class_sections': class_sections,
        'academic_years': academic_years,
    })

# Edit Student
@login_required
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated successfully!')
            return redirect('student_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm(instance=student)
    
    class_sections = ClassSection.objects.all()
    academic_years = AcademicYear.objects.all()
    
    return render(request, 'student_add.html', {
        'form': form,
        'edit': True,
        'class_sections': class_sections,
        'academic_years': academic_years,
    })

# [Keep all other view functions the same as in your original code]

# Delete Student
from django.db import transaction
from django.db.models.deletion import ProtectedError
from django.db import ProgrammingError, DatabaseError
from django.views.decorators.http import require_POST

@login_required
@require_POST
def delete_student(request, pk):
    student = get_object_or_404(Student, pk=pk)

    try:
        with transaction.atomic():
            deleted_count, per_model = student.delete()   # <- returns (#, {app.Model: #})
        messages.success(
            request,
            f"Deleted {deleted_count} row(s). Details: { {m.__name__: n for m, n in per_model.items()} }"
        )
    except ProtectedError as e:
        messages.error(request, f"Cannot delete: related records protect this student: {e.protected_objects}")
    except (ProgrammingError, DatabaseError) as e:
        # Do not swallow; show the exact DB error so you know why it's not deleted
        messages.error(request, f"Database error during delete: {e}")
    except Exception as e:
        messages.error(request, f"Unexpected error: {e}")

    return redirect('student_list')  # or 'students:student_list' if namespaced

# List Students
from django.db.models import Q, Count

from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import Q, Count
from django.shortcuts import render
from .models import Student  # and Notification if you want, but not required here

from django.db import connection

from fees.models import StudentFeeAssignment  
from django.db.models import Count, Q, Value, IntegerField
from django.db import connection


def _table_exists(name: str) -> bool:
    try:
        return name in connection.introspection.table_names()
    except Exception:
        return False
    
@login_required
def student_list(request):
    students = (
        Student.objects
        .select_related('class_section', 'academic_year')
    )

    # Only annotate if the notifications table exists
    if _table_exists('students_notification'):
        students = students.annotate(
            unread_count=Count('notifications', filter=Q(notifications__is_read=False))
        )
    else:
        # Safe fallback so the template never breaks
        students = students.annotate(
            unread_count=Value(0, output_field=IntegerField())
        )

    return render(request, 'student_list.html', {'students': students})

from django.contrib import messages
from django.shortcuts import get_object_or_404, render, redirect
from .models import Student

from django.contrib.contenttypes.models import ContentType
from fees.models import StudentFeeAssignment
# at top of file
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.shortcuts import get_object_or_404, render
from django.contrib.contenttypes.models import ContentType

from .models import Student
from fees.models import StudentFeeAssignment
# students/views.py  (only the notifications parts shown here)
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.db import connection
from django.contrib.contenttypes.models import ContentType

from .models import Student
from fees.models import StudentFeeAssignment

def _table_exists(name: str) -> bool:
    try:
        return name in connection.introspection.table_names()
    except Exception:
        return False

@login_required
def student_notifications(request, pk):
    student = get_object_or_404(Student, pk=pk)

    # Notifications (safe) — latest first
    if _table_exists('students_notification'):
        notes = list(student.notifications.all().order_by('-created_at'))
    else:
        notes = []

    # Attach related fee assignment when present (optional)
    assignment_ct_id = ContentType.objects.get_for_model(StudentFeeAssignment).id
    ids = [
        int(n.object_id) for n in notes
        if n.content_type_id == assignment_ct_id and (n.object_id or "").isdigit()
    ]
    assignments_map = {
        a.id: a
        for a in StudentFeeAssignment.objects
            .filter(id__in=ids)
            .select_related('fee_structure', 'fee_structure__fee_category')
    }
    for n in notes:
        n.assignment = None
        if n.content_type_id == assignment_ct_id and (n.object_id or "").isdigit():
            n.assignment = assignments_map.get(int(n.object_id))

    # Also build per-student summary (optional)
    assignments = (
        StudentFeeAssignment.objects
        .filter(student=student)
        .select_related('fee_structure', 'fee_structure__fee_category')
        .order_by('fee_structure__fee_category__name', 'id')
    )

    return render(request, 'student_notifications.html', {
        'student': student,
        'notifications': notes,
        'assignments': assignments,
        'focus_id': request.GET.get('focus'),   # <-- id of just-created notification
    })

@login_required
def mark_notifications_read(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        if _table_exists('students_notification'):
            student.notifications.filter(is_read=False).update(is_read=True)
            messages.success(request, f"Marked notifications as read for {student.name}.")
        else:
            messages.error(request, "Notifications table not found. Run migrations first.")
    return redirect('student_notifications', pk=student.pk)

# Search Students
@login_required
def search_student(request):
    query = request.GET.get('q', '')
    students = Student.objects.filter(
        Q(name__icontains=query) | Q(admission_number__icontains=query)
    )
    return render(request, 'student_list.html', {
        'students': students,
        'search_query': query
    })

# Export to Excel
@login_required
def export_students_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "ID", "Name", "Admission Number", "Roll No", "Class-Section", "Academic Year", "Gender",
        "DOB", "Father", "Mother", "Guardian", "Guardian Relation", "Aadhaar", "Religion", "Community",
        "Blood Group", "Status", "Contact", "Email", "Address", "RFID", "Student Type",
        "Transport Mode", "Pickup Point", "Bus Route", "Hostel Name", "Room Number"
    ]
    ws.append(headers)

    for student in Student.objects.all():
        ws.append([
            student.id,
            student.name,
            student.admission_number,
            student.roll_no,
            str(student.class_section) if student.class_section else '',
            str(student.academic_year) if student.academic_year else '',
            student.gender,
            student.dob.strftime('%Y-%m-%d') if student.dob else '',
            student.father_name,
            student.mother_name,
            student.guardian_name,
            student.guardian_relation,
            student.aadhaar_number,
            student.religion,
            student.community,
            student.blood_group,
            student.status,
            student.contact_number,
            student.email,
            student.address,
            student.rfid_code,
            student.student_category,
            student.transport_mode,
            student.pickup_point,
            student.bus_route_number,
            student.hostel_name,
            student.room_number,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    return response

# Export to PDF
@login_required
def export_students_pdf(request):
    students = Student.objects.all()
    template = get_template('student_pdf_template.html')
    html_content = template.render({'students': students})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="students.pdf"'

    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as output:
        HTML(string=html_content).write_pdf(target=output.name)
        output.seek(0)
        response.write(output.read())

    return response

# views.py
import openpyxl
from datetime import datetime, date
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render

CHOICE_SETS = {
    "gender": {"M", "F"},
    "religion": {"Hindu", "Christian", "Muslim"},
    "community": {"OC", "BC", "MBC", "SC", "ST", "FC"},
    "status": {"Active", "Alumni", "Left"},
    "student_category": {"day_scholar", "hosteller"},
    "transport_mode": {"school_bus", "private_van", "auto", "own_vehicle", "walk", "other", "", None},
}

def _as_date(v):
    if not v:
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip().split()[0]
    return datetime.strptime(s, "%Y-%m-%d").date()

@login_required
def upload_students_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            created, updated, errors = 0, 0, 0
            error_messages = []

            with transaction.atomic():
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not row or not any(row):
                            continue

                        # Read by column index (keeps your current template compatible)
                        name = (row[0] or "").strip()
                        if not name:
                            raise ValueError("Name is required")

                        gender = (row[1] or "").strip()
                        if gender not in CHOICE_SETS["gender"]:
                            raise ValueError(f"Invalid gender '{gender}'")

                        dob = _as_date(row[2])

                        # Parse class-section like "7 - A"
                        class_section_str = str(row[3]).strip()
                        try:
                            class_name, section = map(str.strip, class_section_str.split('-'))
                        except Exception:
                            raise ValueError(f"Invalid class-section format: '{class_section_str}'")
                        class_section = ClassSection.objects.filter(
                            class_name=class_name, section=section
                        ).first()
                        if not class_section:
                            raise ValueError(f"Class Section '{class_section_str}' not found")

                        # Academic year must exist (or you can auto-create)
                        academic_year_str = str(row[4]).strip()
                        ay = AcademicYear.objects.filter(name=academic_year_str).first()
                        if not ay:
                            raise ValueError(f"Academic Year '{academic_year_str}' not found")

                        religion = (row[13] or "").strip()
                        if religion and religion not in CHOICE_SETS["religion"]:
                            raise ValueError(f"Invalid religion '{religion}'")

                        community = (row[14] or "").strip()
                        if community and community not in CHOICE_SETS["community"]:
                            raise ValueError(f"Invalid community '{community}'")

                        status = (row[19] or "").strip()
                        if status and status not in CHOICE_SETS["status"]:
                            raise ValueError(f"Invalid status '{status}'")

                        student_category = (row[22] or "").strip()
                        if student_category and student_category not in CHOICE_SETS["student_category"]:
                            raise ValueError(f"Invalid student_category '{student_category}'")

                        transport_mode = (row[23] or "").strip()
                        if transport_mode not in CHOICE_SETS["transport_mode"]:
                            raise ValueError(f"Invalid transport_mode '{transport_mode}'")

                        # Unpack the rest
                        defaults = {
                            'name': name,
                            'gender': gender,
                            'dob': dob,
                            'class_section': class_section,
                            'academic_year': ay,
                            'father_name': row[5],
                            'mother_name': row[6],
                            'guardian_name': row[7],
                            'guardian_relation': row[8],
                            'address': row[9],
                            'contact_number': row[10],
                            'email': row[11],
                            'aadhaar_number': row[12],
                            'religion': religion or "",
                            'community': community or "",
                            'blood_group': row[15],
                            'previous_school': row[16],
                            'transfer_certificate_number': row[17],
                            'date_of_leaving': _as_date(row[18]),
                            'status': status or "Active",
                            'rfid_code': row[21],
                            'student_category': student_category or "day_scholar",
                            'transport_mode': transport_mode or None,
                            'pickup_point': row[24],
                            'bus_route_number': row[25],
                            'hostel_name': row[26],
                            'room_number': row[27],
                            'biometric_id': row[28],
                        }

                        admission_number = (row[20] or "").strip()
                        aadhaar = (row[12] or "").strip()
                        biometric = (row[28] or "").strip()

                        # Preferred key order: admission_number > biometric_id > aadhaar_number
                        lookup_kwargs = {}
                        if admission_number:
                            lookup_kwargs = {"admission_number": admission_number}
                        elif biometric:
                            lookup_kwargs = {"biometric_id": biometric}
                        elif aadhaar:
                            lookup_kwargs = {"aadhaar_number": aadhaar}
                        else:
                            # As a last resort, use name+dob (not guaranteed unique!)
                            if not dob:
                                raise ValueError("Need at least one identifier: admission_number/biometric_id/aadhaar_number or (name+dob)")
                            lookup_kwargs = {"name": name, "dob": dob}

                        student, created_flag = Student.objects.update_or_create(
                            **lookup_kwargs,
                            defaults=defaults
                        )

                        # If admission_number was blank, your Student.save() will auto-generate it now
                        if created_flag:
                            created += 1
                        else:
                            updated += 1

                    except Exception as e:
                        error_messages.append(f"Row {i}: {str(e)}")
                        errors += 1
                        continue

            if created or updated:
                messages.success(request, f"Processed {created} created, {updated} updated.")
            if errors:
                messages.warning(request, f"Encountered {errors} row error(s).")
                for err in error_messages[:5]:
                    messages.error(request, err)
                if errors > 5:
                    messages.info(request, f"...plus {errors-5} more error(s)")

        except Exception as e:
            messages.error(request, f"Failed to process file: {e}")

        return redirect('student_list')

    return render(request, 'upload_students_excel.html')

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
import openpyxl
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from datetime import datetime
import os
from .models import Staff
from .forms import StaffForm

# Permission functions
def is_admin(user):
    return hasattr(user, 'role') and user.role == 'admin'

def is_hr(user):
    return hasattr(user, 'staff') and user.staff.staff_role == 'HR'

# Staff List
def staff_list(request):
    query = request.GET.get('q', '')
    selected_category = request.GET.get('staff_category', '')
    selected_department = request.GET.get('department', '')
    selected_designation = request.GET.get('designation', '')

    staff = Staff.objects.all()

    if query:
        staff = staff.filter(
            Q(full_name__icontains=query) |
            Q(staff_id__icontains=query) |
            Q(contact_number__icontains=query)
        )
    if selected_category:
        staff = staff.filter(staff_category=selected_category)
    if selected_department:
        staff = staff.filter(department=selected_department)
    if selected_designation:
        staff = staff.filter(designation=selected_designation)

    context = {
        'staff': staff,
        'query': query,
        'categories': Staff.STAFF_CATEGORY_CHOICES,
        'departments': Staff.DEPARTMENT_CHOICES,
        'designations': Staff.DESIGNATION_CHOICES,
        'selected_category': selected_category,
        'selected_department': selected_department,
        'selected_designation': selected_designation,
    }
    return render(request, 'staff/staff_list.html', context)

# Add Staff
@login_required
@user_passes_test(lambda u: is_admin(u) or is_hr(u))
def add_staff(request):
    if request.method == 'POST':
        form = StaffForm(request.POST, request.FILES)
        if form.is_valid():
            staff = form.save(commit=False)
            cropped_data = request.POST.get('cropped_image')
            if cropped_data:
                try:
                    format, imgstr = cropped_data.split(';base64,')
                    ext = format.split('/')[-1]
                    staff.photo = ContentFile(base64.b64decode(imgstr), name=f"{staff.full_name.replace(' ', '_')}_photo.{ext}")
                except Exception as e:
                    print("Error processing cropped image:", e)
            staff.save()
            messages.success(request, "✅ Staff added successfully.")
            return redirect('staff_list')
        else:
            messages.error(request, "Please correct the errors.")
    else:
        form = StaffForm()
    return render(request, 'staff/staff_add.html', {'form': form})

# Edit Staff
@login_required
@user_passes_test(lambda u: is_admin(u) or is_hr(u))
def edit_staff(request, pk):
    staff = get_object_or_404(Staff, pk=pk)
    form = StaffForm(request.POST or None, request.FILES or None, instance=staff)
    if request.method == 'POST' and form.is_valid():
        staff = form.save(commit=False)
        cropped_data = request.POST.get('cropped_image')
        if cropped_data:
            try:
                format, imgstr = cropped_data.split(';base64,')
                ext = format.split('/')[-1]
                staff.photo = ContentFile(base64.b64decode(imgstr), name=f"{staff.full_name.replace(' ', '_')}_photo.{ext}")
            except Exception as e:
                print("Error cropping image in edit:", e)
        staff.save()
        messages.success(request, "✅ Staff updated successfully.")
        return redirect('staff_list')
    return render(request, 'staff/staff_edit.html', {'form': form, 'staff': staff})

# Delete Staff
@login_required
@user_passes_test(is_admin)
def delete_staff(request, pk):
    staff = get_object_or_404(Staff, pk=pk)
    staff.delete()
    messages.success(request, "Staff record deleted successfully.")
    return redirect('staff_list')

# Export to PDF
@login_required
def export_staff_pdf(request):
    staff = Staff.objects.all()
    html_string = render_to_string('staff/staff_pdf_template.html', {'staff': staff})
    html = HTML(string=html_string)

    fd, path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    html.write_pdf(target=path)

    with open(path, 'rb') as pdf_file:
        pdf_data = pdf_file.read()
    os.remove(path)

    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="staff_report.pdf"'
    return response

# Export to Excel
@login_required
def export_staff_excel(request):
    staff = Staff.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Staff List"

    headers = ['Staff ID', 'Name', 'Category', 'Department', 'Designation', 'Contact', 'Status']
    sheet.append(headers)

    for s in staff:
        sheet.append([
            s.staff_id,
            s.full_name,
            s.get_staff_category_display(),
            s.get_department_display(),
            s.get_designation_display(),
            s.contact_number,
            s.get_status_display()
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=staff_list.xlsx'
    workbook.save(response)
    return response

# Upload Excel
from django.db import transaction, IntegrityError
from datetime import datetime, date, timedelta

@login_required
def upload_staff_excel_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        created = updated = errors = 0
        error_messages = []

        # Build quick lookup dicts from model choices (code -> display)
        DEPT_CODES = {c[0]: c[1] for c in Staff.DEPARTMENT_CHOICES}
        DESIG_CODES = {c[0]: c[1] for c in Staff.DESIGNATION_CHOICES}
        GENDER_CODES = dict(Staff.GENDER_CHOICES)

        # Helpers
        def excel_serial_to_date(n):
            # Excel's 1900 date system (with 1900 leap-year bug)
            # date = 1899-12-30 + n days
            try:
                return (datetime(1899, 12, 30) + timedelta(days=int(n))).date()
            except Exception:
                return None

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, date):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, (int, float)):
                # Likely an Excel serial number
                return excel_serial_to_date(val)
            if isinstance(val, str):
                s = val.strip().split()[0]
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except ValueError:
                        continue
            return None

        def normalize_choice(raw, valid_codes, default_code):
            """
            Accepts either the internal code (e.g., 'SCIENCE') or the display label (e.g., 'Science'),
            or a close uppercase token; falls back to default_code.
            """
            if not raw:
                return default_code
            s = str(raw).strip()
            # direct exact code
            if s in valid_codes:
                return s
            # match by display (case-insensitive)
            for code, label in valid_codes.items():
                if s.lower() == str(label).lower():
                    return code
            # try a loose prefix match against code
            s_up = s.upper()
            for code in valid_codes:
                if s_up.startswith(code[:3]):
                    return code
            return default_code

        def normalize_gender(raw):
            if not raw:
                return None
            s = str(raw).strip().lower()
            if s in ('m', 'male'):
                return 'M'
            if s in ('f', 'female'):
                return 'F'
            if s in ('o', 'other', 'others'):
                return 'O'
            # if it's already a valid code
            if raw in GENDER_CODES:
                return raw
            return None

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            # Expected columns (0-based):
            # 0: staff_id (optional)
            # 1: full_name
            # 2: gender
            # 3: dob
            # 4: doj
            # 5: contact_number
            # 6: email
            # 7: address
            # 8: employment_type
            # 9: salary
            # 10: status
            # 11: remarks
            # 12: staff_category
            # 13: staff_role
            # 14: department
            # 15: designation
            # 16: rfid_code
            # 17: biometric_id

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not any(row):
                        continue

                    # Safe access helper
                    def col(n, default=None):
                        return row[n] if len(row) > n and row[n] is not None else default

                    # Parse and normalize
                    staff_id_cell = col(0, '').strip() if isinstance(col(0), str) else col(0)
                    full_name = (col(1, '') or '').strip()
                    gender = normalize_gender(col(2))
                    dob = parse_date(col(3))
                    doj = parse_date(col(4)) or timezone.now().date()
                    contact_number = str(col(5, '') or '').strip()
                    email = (col(6, None) or None)
                    address = (col(7, '') or '')
                    employment_type = (col(8, 'Permanent') or 'Permanent')
                    salary_raw = col(9, 0.00)
                    try:
                        salary = float(salary_raw or 0.00)
                    except Exception:
                        salary = 0.00
                    status = (col(10, 'Active') or 'Active')
                    remarks = (col(11, '') or '')
                    staff_category = (col(12, 'TEACHER') or 'TEACHER')
                    staff_role = (col(13, 'TEACHER') or 'TEACHER')
                    department = normalize_choice(col(14, 'OTHERS'), DEPT_CODES, 'OTHERS')
                    designation = normalize_choice(col(15, 'TEACHER'), DESIG_CODES, 'TEACHER')
                    rfid_code = col(16, None)
                    biometric_id = col(17, None)

                    if not full_name:
                        raise ValueError("Full name is required")

                    if not dob:
                        raise ValueError("Invalid date for Date of Birth")

                    if gender is None:
                        raise ValueError("Invalid gender (use M/F/O or Male/Female/Other)")

                    defaults = {
                        'full_name': full_name,
                        'gender': gender,
                        'dob': dob,
                        'doj': doj,
                        'contact_number': contact_number,
                        'email': email,
                        'address': address,
                        'employment_type': employment_type,
                        'salary': salary,
                        'status': status,
                        'remarks': remarks,
                        'staff_category': staff_category,
                        'staff_role': staff_role,
                        'department': department,
                        'designation': designation,
                        'rfid_code': rfid_code,
                        'biometric_id': biometric_id,
                    }

                    with transaction.atomic():
                        if staff_id_cell:
                            # Update by given staff_id if present; create if not exists
                            obj, created_flag = Staff.objects.update_or_create(
                                staff_id=str(staff_id_cell).strip(),
                                defaults=defaults
                            )
                        else:
                            # No staff_id supplied -> create and let model auto-generate
                            obj = Staff.objects.create(**defaults)
                            created_flag = True

                    if created_flag:
                        created += 1
                    else:
                        updated += 1

                except IntegrityError as ie:
                    errors += 1
                    error_messages.append(f"Row {i}: Integrity error (possible duplicate staff_id/rfid/biometric). {ie}")
                except Exception as e:
                    errors += 1
                    error_messages.append(f"Row {i}: {str(e)}")

            if created or updated:
                messages.success(request, f"Processed {created} new and {updated} updated record(s).")
            if errors:
                messages.warning(request, f"Encountered {errors} error(s). Showing a few below.")
                for msg in error_messages[:5]:
                    messages.error(request, msg)

        except Exception as e:
            messages.error(request, f"Failed to process file: {str(e)}")

        return redirect('staff_list')

    return render(request, 'staff/upload_staff_excel.html')

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Staff  # Import your Staff model

def delete_staff(request, pk):
    staff = get_object_or_404(Staff, pk=pk)

    if request.method == "POST":
        staff.delete()
        messages.success(request, "Staff record deleted successfully.")
        return redirect('staff_list')  # Go back to the list page

    return redirect('staff_list')
